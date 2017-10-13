'''
Created on Oct 6, 2017

@author: djames
'''

from openpyxl import load_workbook

from openpyxl.styles import Alignment, NamedStyle, Font, Border, Side, PatternFill

from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles.differential import DifferentialStyle


# Auto Width
# https://stackoverflow.com/questions/13197574/openpyxl-adjust-column-width-size

# Hide:
# https://stackoverflow.com/questions/38527725/how-can-i-hide-columns-in-openpyxl
# ws.column_dimensions.group(start='B', end='CU', hidden=True)


###
# from openpyxl.styles import NamedStyle, Font, Border, Side
# highlight = NamedStyle(name="highlight")
# highlight.font = Font(bold=True, size=20)
# bd = Side(style='thick', color="000000")
# highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)


THIN_BORDER = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
RED_FILL = PatternFill(start_color='FF9999',
                       end_color='FF9999',
                       fill_type='solid')
GREEN_FILL = PatternFill(start_color='CCFF99',
                         end_color='CCFF99',
                         fill_type='solid')


VERT_STYLE = NamedStyle(name='VERT_STYLE')  # Style for parameters that shoudl be rotated
VERT_STYLE.alignment = Alignment(textRotation=90, vertical='bottom', horizontal='center')
VERT_STYLE.font = Font(bold=True)
VERT_STYLE.border = THIN_BORDER

VERT2 = DifferentialStyle()
VERT2.alignment = Alignment(textRotation=90, vertical='bottom', horizontal='center')

# List of sheets in run report and compare that should rotate line 2.
VERT_ROWS_SHEETS = ['Result', 'Results', 'CommonParamFlat',
                    'Results Exceeding Tolerance', 'aleph', 'bet']


def set_row_style(sheet, row, style):
    """
    apply the given style to an entire row
    :param sheet: openpyxl sheet
    :param row: int, row number
    :param style: openpyxl styel
    """

    for cell in sheet[row]:
        cell.style = style


def col_match(sheet, row, value):
    """
    find all columns matching value

    Return first matching column
    :param sheet: an openpyxl sheet
    :param row:   row descriptor (e.g. '2:2')
    :param value: the value to look for.
    """
    for cell in sheet[row]:
        if cell.value == value:
            return cell.column


def auto_width(sheet):
    """
    Adjust the column width, mimicking Excel's auto function

    Must approximate for variable width columns
    :param sheet: an openpyxl sheet object.
    """
    column_widths = []
    for row in sheet.iter_rows():
        for i, cell in enumerate(row):
            try:
                # if cell alignment is vertical, use 4, else len(str(cell.value)
                if cell.alignment.textRotation == 90:
                    cell_w = 4
                else:
                    cell_w = len(str(cell.value))
                column_widths[i] = max(column_widths[i], cell_w)
            except IndexError:
                column_widths.append(len(str(cell.value)))

    for i, column_width in enumerate(column_widths):
        # 1.4 is a fudge factor for variable width fonts
        sheet.column_dimensions[get_column_letter(i + 1)].width = column_width * 1.4


def main():
    '''Main'''

    my_wb = load_workbook(filename='51-SysTstDev-192-SysTstDev-5.xlsx')
    print(col_match(my_wb['Results'], '2:2', 'diff'))
    tclm = col_match(my_wb['Results'], '2:2', 'diff')
    thresh_column = tclm + ':' + tclm
    for sheet in VERT_ROWS_SHEETS:
        try:
            set_row_style(my_wb[sheet], '2:2', VERT_STYLE)
        except KeyError:
            pass
    for sheet in my_wb.get_sheet_names():
        auto_width(my_wb[sheet])

    my_wb['Results'].column_dimensions.group(start='K', end='Q', hidden=True)

    my_wb['Results'].conditional_formatting.add(thresh_column,
                                                CellIsRule(operator='greaterThan',
                                                           formula=['5'],
                                                           fill=GREEN_FILL,
                                                           stopIfTrue=True))
    my_wb['Results'].conditional_formatting.add('AZ:AZ',
                                                CellIsRule(operator='lessThan',
                                                           formula=['-5'],
                                                           fill=RED_FILL,
                                                           stopIfTrue=True))
    my_wb.active = my_wb.index(my_wb['Results'])  # TODO Result for report
    my_wb.save(filename='51-SysTstDev-192-SysTstDev-5PLUS.xlsx')

if __name__ == '__main__':
    main()
